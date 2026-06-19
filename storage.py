"""Excel storage for receipts (openpyxl).

One isolated folder per Telegram user, and **one workbook per calendar year**
inside it, so each year's receipts live in their own file and never mix:

  data/users/<telegram_user_id>/receipts_<year>.xlsx
  data/users/<telegram_user_id>/photos/
  data/users/<telegram_user_id>/.reported_<year>.flag   (year-end report sent)

Each workbook has:
  Receipts sheet — one row per receipt.
  Summary sheet  — that year's totals via live Excel formulas.

Receipt IDs are unique per user (global, not reset each year) so /fix and
/delete by ID are never ambiguous across files.
"""

import glob
import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

HEADERS = ["ID", "Added", "Receipt Date", "Year", "Store",
           "Subtotal", "Sales Tax", "Total", "Photo", "Hash", "User"]
MONEY_COLS = (6, 7, 8)          # Subtotal, Sales Tax, Total
COL_ID, COL_YEAR, COL_STORE = 1, 4, 5
COL_SUBTOTAL, COL_TAX, COL_TOTAL = 6, 7, 8
COL_DATE, COL_HASH = 3, 10
MONEY_FMT = '"$"#,##0.00'

_YEAR_FILE = re.compile(r"receipts_(\d{4})\.xlsx$")


# ----------------------------------------------------------------- paths

def user_dir(user_id) -> str:
    return os.path.join(DATA_DIR, "users", str(user_id))


def photos_dir(user_id) -> str:
    return os.path.join(user_dir(user_id), "photos")


def excel_path(user_id, year: int) -> str:
    return os.path.join(user_dir(user_id), f"receipts_{year}.xlsx")


def _year_files(user_id) -> dict[int, str]:
    """{year: path} for every per-year workbook this user has."""
    out = {}
    for p in glob.glob(os.path.join(user_dir(user_id), "receipts_*.xlsx")):
        m = _YEAR_FILE.search(os.path.basename(p))
        if m:
            out[int(m.group(1))] = p
    return dict(sorted(out.items()))


def all_user_ids() -> list[int]:
    base = os.path.join(DATA_DIR, "users")
    if not os.path.isdir(base):
        return []
    return sorted(int(n) for n in os.listdir(base)
                  if n.isdigit() and os.path.isdir(os.path.join(base, n)))


# ----------------------------------------------------------------- workbook

def _open(user_id, year: int) -> Workbook:
    os.makedirs(photos_dir(user_id), exist_ok=True)
    path = excel_path(user_id, year)
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [6, 18, 12, 8, 24, 11, 11, 11, 28, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.create_sheet("Summary")
    return wb


def _data_rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def _has_data(ws) -> bool:
    return any(r[COL_ID - 1] is not None for r in _data_rows(ws))


def _rebuild_summary(wb: Workbook) -> None:
    ws = wb["Receipts"]
    sm = wb["Summary"]
    sm.delete_rows(1, sm.max_row)

    years = sorted({row[COL_YEAR - 1] for row in _data_rows(ws) if row[COL_YEAR - 1]})
    sm.append(["Year", "Receipts", "Total Spent", "Total Sales Tax"])
    for cell in sm[1]:
        cell.font = Font(bold=True)
    for w, col in zip((10, 10, 14, 16), "ABCD"):
        sm.column_dimensions[col].width = w

    yc = get_column_letter(COL_YEAR)
    tc = get_column_letter(COL_TOTAL)
    xc = get_column_letter(COL_TAX)
    for y in years:
        sm.append([
            y,
            f'=COUNTIF(Receipts!{yc}:{yc},A{sm.max_row + 1})',
            f'=SUMIF(Receipts!{yc}:{yc},A{sm.max_row + 1},Receipts!{tc}:{tc})',
            f'=SUMIF(Receipts!{yc}:{yc},A{sm.max_row + 1},Receipts!{xc}:{xc})',
        ])
    last = sm.max_row
    sm.append(["TOTAL",
               f"=SUM(B2:B{last})", f"=SUM(C2:C{last})", f"=SUM(D2:D{last})"])
    for cell in sm[sm.max_row]:
        cell.font = Font(bold=True)
    for row in sm.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = MONEY_FMT


def _save(wb: Workbook, user_id, year: int) -> None:
    try:
        wb.save(excel_path(user_id, year))
    except PermissionError as e:
        raise RuntimeError(
            f"Can't write receipts_{year}.xlsx — close it in Excel and try again."
        ) from e


def _append(user_id, year: int, values: list) -> None:
    """Append one fully-formed row (used by add_receipt, migration, moves)."""
    wb = _open(user_id, year)
    ws = wb["Receipts"]
    ws.append(values)
    for col in MONEY_COLS:
        ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
    _rebuild_summary(wb)
    _save(wb, user_id, year)


# ----------------------------------------------------------------- helpers

def _year_from_date(date: str | None) -> int:
    if date and len(str(date)) >= 4 and str(date)[-4:].isdigit():
        return int(str(date)[-4:])
    return datetime.now().year


def _iter_all_rows(user_id):
    """Yield every receipt row (values tuple) across all of a user's files."""
    for path in _year_files(user_id).values():
        ws = load_workbook(path)["Receipts"]
        for row in _data_rows(ws):
            if row[COL_ID - 1] is not None:
                yield row


def _all_rows_sorted(user_id) -> list:
    return sorted(_iter_all_rows(user_id), key=lambda r: r[COL_ID - 1])


def _next_id(user_id) -> int:
    ids = [r[COL_ID - 1] for r in _iter_all_rows(user_id)]
    return (max(ids) + 1) if ids else 1


def years_with_data(user_id) -> list[int]:
    out = []
    for year, path in _year_files(user_id).items():
        if _has_data(load_workbook(path)["Receipts"]):
            out.append(year)
    return sorted(out)


# ----------------------------------------------------------------- CRUD

def find_duplicate(user_id, *, file_hash: str = "", store: str = "",
                   date: str = "", tax=None, total=None) -> tuple[int, str] | None:
    """Return (receipt_id, reason) if this receipt is already stored for this
    user (searches every year's file). Matches the exact same photo (hash) or
    the same parsed receipt (store + date + tax + total)."""
    store_key = (store or "").strip().lower()
    for row in _iter_all_rows(user_id):
        if file_hash and row[COL_HASH - 1] == file_hash:
            return row[COL_ID - 1], "same photo"
        if (total is not None and store_key
                and (row[COL_STORE - 1] or "").strip().lower() == store_key
                and (row[COL_DATE - 1] or "") == (date or "")
                and row[COL_TOTAL - 1] == total
                and row[COL_TAX - 1] == tax):
            return row[COL_ID - 1], "same store, date and amounts"
    return None


def add_receipt(user_id, *, store: str, date: str, subtotal, tax, total,
                photo: str, user: str, file_hash: str = "") -> int:
    """Append a receipt into the workbook for its year; returns its ID."""
    year = _year_from_date(date)
    rid = _next_id(user_id)
    values = [rid, datetime.now().strftime("%m/%d/%Y %H:%M"), date or "", year,
              store or "Unknown", subtotal, tax, total, photo, file_hash, user]
    _append(user_id, year, values)
    return rid


def _locate(user_id, rid: int):
    """Return (year, wb, ws, row_number) for a receipt id, or None."""
    for year, path in _year_files(user_id).items():
        wb = load_workbook(path)
        ws = wb["Receipts"]
        for row in ws.iter_rows(min_row=2):
            if row[COL_ID - 1].value == rid:
                return year, wb, ws, row[COL_ID - 1].row
    return None


def update_receipt(user_id, rid: int, field: str, value) -> bool:
    """field: 'tax' | 'total' | 'subtotal' | 'store' | 'date'.
    A date edit that changes the year moves the receipt to the right file."""
    col = {"tax": COL_TAX, "total": COL_TOTAL, "subtotal": COL_SUBTOTAL,
           "store": COL_STORE, "date": COL_DATE}[field]
    loc = _locate(user_id, rid)
    if not loc:
        return False
    year, wb, ws, rownum = loc

    if field == "date":
        new_year = _year_from_date(value)
        if new_year != year:
            # Move the row to the correct year's file, keeping its ID.
            vals = [c.value for c in ws[rownum]]
            vals[COL_DATE - 1] = value
            vals[COL_YEAR - 1] = new_year
            ws.delete_rows(rownum, 1)
            _finish_after_edit(wb, ws, user_id, year)
            _append(user_id, new_year, vals)
            return True
        ws.cell(row=rownum, column=COL_DATE).value = value
        ws.cell(row=rownum, column=COL_YEAR).value = new_year
    else:
        ws.cell(row=rownum, column=col).value = value
        if col in MONEY_COLS:
            ws.cell(row=rownum, column=col).number_format = MONEY_FMT

    _finish_after_edit(wb, ws, user_id, year)
    return True


def _finish_after_edit(wb, ws, user_id, year) -> None:
    """Save (or delete) a workbook after a row was changed or removed."""
    if _has_data(ws):
        _rebuild_summary(wb)
        _save(wb, user_id, year)
    else:
        try:
            os.remove(excel_path(user_id, year))
        except OSError:
            _rebuild_summary(wb)
            _save(wb, user_id, year)


def delete_receipt(user_id, rid: int) -> bool:
    loc = _locate(user_id, rid)
    if not loc:
        return False
    year, wb, ws, rownum = loc
    ws.delete_rows(rownum, 1)
    _finish_after_edit(wb, ws, user_id, year)
    return True


def last_receipt_id(user_id) -> int | None:
    ids = [r[COL_ID - 1] for r in _iter_all_rows(user_id)]
    return max(ids) if ids else None


def get_receipt(user_id, rid: int) -> dict | None:
    for row in _iter_all_rows(user_id):
        if row[COL_ID - 1] == rid:
            return dict(zip(HEADERS, row))
    return None


def get_summary(user_id, year: int | None = None) -> dict:
    """Totals for one year, or across all years, plus per-year breakdown."""
    if year is not None:
        path = _year_files(user_id).get(year)
        rows = _data_rows(load_workbook(path)["Receipts"]) if path else []
        rows = [r for r in rows if r[COL_ID - 1] is not None]
    else:
        rows = list(_iter_all_rows(user_id))

    per_year: dict[int, dict] = {}
    for r in rows:
        y = r[COL_YEAR - 1]
        d = per_year.setdefault(y, {"count": 0, "spent": 0.0, "tax": 0.0})
        d["count"] += 1
        d["spent"] += r[COL_TOTAL - 1] or 0
        d["tax"] += r[COL_TAX - 1] or 0
    return {
        "count": len(rows),
        "spent": round(sum(r[COL_TOTAL - 1] or 0 for r in rows), 2),
        "tax": round(sum(r[COL_TAX - 1] or 0 for r in rows), 2),
        "per_year": {y: {k: round(v, 2) if isinstance(v, float) else v
                         for k, v in d.items()}
                     for y, d in sorted(per_year.items())},
    }


def get_last(user_id, n: int = 5) -> list[dict]:
    rows = _all_rows_sorted(user_id)[-n:]
    return [dict(zip(HEADERS, r)) for r in rows]


def excel_exists(user_id, year: int | None = None) -> bool:
    if year is not None:
        return year in years_with_data(user_id)
    return bool(years_with_data(user_id))


# ------------------------------------------------- year-end report markers

def _flag_path(user_id, year: int) -> str:
    return os.path.join(user_dir(user_id), f".reported_{year}.flag")


def is_reported(user_id, year: int) -> bool:
    return os.path.exists(_flag_path(user_id, year))


def mark_reported(user_id, year: int) -> None:
    os.makedirs(user_dir(user_id), exist_ok=True)
    with open(_flag_path(user_id, year), "w", encoding="utf-8") as f:
        f.write(datetime.now().isoformat())


# ------------------------------------------------- legacy migration

def migrate_legacy(user_id) -> bool:
    """Split an old single-file receipts.xlsx into per-year workbooks.
    Runs once; the old file is renamed to a backup. Returns True if migrated."""
    legacy = os.path.join(user_dir(user_id), "receipts.xlsx")
    if not os.path.exists(legacy):
        return False
    ws = load_workbook(legacy)["Receipts"]
    for row in _data_rows(ws):
        if row[COL_ID - 1] is None:
            continue
        vals = list(row)
        if len(vals) == 10:        # pre-Hash layout: insert empty Hash column
            vals = vals[:9] + ["", vals[9]]
        while len(vals) < len(HEADERS):
            vals.append("")
        year = vals[COL_YEAR - 1] or _year_from_date(vals[COL_DATE - 1])
        vals[COL_YEAR - 1] = year
        _append(user_id, year, vals)
    os.replace(legacy, os.path.join(user_dir(user_id), "receipts.pre-split-backup.xlsx"))
    # Don't fire year-end reports for years that already completed before upgrade.
    current = datetime.now().year
    for y in years_with_data(user_id):
        if y < current:
            mark_reported(user_id, y)
    return True


def migrate_all() -> int:
    return sum(1 for uid in all_user_ids() if migrate_legacy(uid))
